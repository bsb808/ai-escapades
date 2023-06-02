#!/usr/bin/env python3
import os
import openai
import openpyxl 
from dotenv import load_dotenv

# Load the .env variables and set API key.
load_dotenv() 
openai.api_key = os.getenv("OPENAI_API_KEY")

# Open Excel file with authors and abstracts.
datafile='budget.xlsx'
datafile='MAEBudgetRequirementInputs_nonlabor_FY24t.xlsx'
wb = openpyxl.load_workbook(datafile)
ws = wb["FY24"]

cnt=0
for row in ws.iter_rows(
        min_row = 5, max_row = ws.max_row,
        min_col=1, max_col=ws.max_column):
    category = row[3].value
    requirement = row[5].value
    cost = row[7].value

    prompt = ("""Generate a Necessity Statement to justify a budget item for an engineering university teaching laboratory.

Here is an example.

Category: Office Supplies
Cost: $9000
Requirement: office supplies - estimate based on historical execution
Neccessity Statement: Office supplies for day-to-day business and instruciton of studnets to include printing of course material , whiteboard markers, pens and stationary for notes and correspondence, etc.

Generate a new Necessity Statement based on the following information.

Category: %s
Cost: $%d
Requirement: %s
Neccesity Statement:
    """%(category, cost, requirement))
    #print(prompt)
    cnt+=1
    print("Generating %d statement..."%cnt)
    r = openai.Completion.create(model = "text-davinci-003",
                                 prompt = prompt,
                                 max_tokens = 1000,
                                 temperature=0.6)
    resp = r.choices[0].text
    print(r.choices[0].finish_reason)
    print(resp)
    ws.cell(row=row[0].row, column=7).value = resp

wb.save('budget_justified.xlsx')
    
    
